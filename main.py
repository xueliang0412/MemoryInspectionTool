# 1. 优化界面统计信息模块
# 2. 增加版本号显示
# 3. 统计信息保存至内存监控报告中

import psutil
import pandas as pd
import time
import threading
import os
import matplotlib.pyplot as plt
from datetime import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib
from io import BytesIO
import matplotlib.dates as mdates
import numpy as np

matplotlib.use('TkAgg')
# ---------------------- 添加字体配置 ----------------------
plt.rcParams["font.family"] = ["SimHei", "Microsoft YaHei"]  # 解决中文显示问题
plt.rcParams["axes.unicode_minus"] = False  # 解决负号显示问题


class MemoryMonitorApp:
    def __init__(self, root):
        self.root = root
        # ---------------------- 修改1：添加版本号v1.0 ----------------------
        self.root.title("多进程内存监控工具 v1.0")
        self.root.geometry("1000x800")
        self.root.resizable(True, True)

        # 初始化变量
        self.monitoring = False
        self.monitor_thread = None
        self.process_data = {}  # {进程名: DataFrame}
        self.selected_processes = set()
        self.merge_processes = set()
        self.save_path = os.getcwd()

        # 统计信息相关变量
        self.stats_tree = None

        # 创建UI界面
        self._create_widgets()

    def _create_widgets(self):
        """创建UI组件"""
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 1. 进程选择区域
        process_frame = ttk.LabelFrame(main_frame, text="进程选择", padding="10")
        process_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Label(process_frame, text="可用进程:").grid(row=0, column=0, sticky=tk.W)
        self.process_listbox = tk.Listbox(process_frame, selectmode=tk.EXTENDED, height=6, width=50)
        self.process_listbox.grid(row=1, column=0, padx=(0, 10), pady=5)

        scrollbar = ttk.Scrollbar(process_frame, orient=tk.VERTICAL, command=self.process_listbox.yview)
        scrollbar.grid(row=1, column=1, sticky=tk.NS)
        self.process_listbox.config(yscrollcommand=scrollbar.set)

        btn_frame = ttk.Frame(process_frame)
        btn_frame.grid(row=1, column=2, padx=10)
        ttk.Button(btn_frame, text="刷新进程", command=self._refresh_processes).pack(fill=tk.X, pady=5)
        ttk.Button(btn_frame, text="添加监控", command=self._add_monitor).pack(fill=tk.X, pady=5)
        ttk.Button(btn_frame, text="移除监控", command=self._remove_monitor).pack(fill=tk.X, pady=5)

        ttk.Label(process_frame, text="当前监控进程:").grid(row=0, column=3, sticky=tk.W)
        self.monitor_listbox = tk.Listbox(process_frame, selectmode=tk.EXTENDED, height=6, width=50)
        self.monitor_listbox.grid(row=1, column=3, padx=(0, 10), pady=5)

        # 2. 参数设置区域
        param_frame = ttk.LabelFrame(main_frame, text="监控参数", padding="10")
        param_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Label(param_frame, text="监控时长:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        self.duration_var = tk.StringVar(value="60")
        ttk.Entry(param_frame, textvariable=self.duration_var, width=10).grid(row=0, column=1, sticky=tk.W, pady=5)
        self.duration_unit = ttk.Combobox(param_frame, values=["秒", "分钟", "小时"], width=6, state="readonly")
        self.duration_unit.current(0)
        self.duration_unit.grid(row=0, column=2, sticky=tk.W, padx=5, pady=5)

        ttk.Label(param_frame, text="采样间隔:").grid(row=0, column=3, sticky=tk.W, padx=5, pady=5)
        self.interval_var = tk.StringVar(value="5")
        ttk.Entry(param_frame, textvariable=self.interval_var, width=10).grid(row=0, column=4, sticky=tk.W, pady=5)
        self.interval_unit = ttk.Combobox(param_frame, values=["秒", "分钟"], width=6, state="readonly")
        self.interval_unit.current(0)
        self.interval_unit.grid(row=0, column=5, sticky=tk.W, padx=5, pady=5)

        ttk.Label(param_frame, text="报告保存路径:").grid(row=0, column=6, sticky=tk.W, padx=5, pady=5)
        self.path_var = tk.StringVar(value=os.getcwd())
        ttk.Entry(param_frame, textvariable=self.path_var, width=30).grid(row=0, column=7, sticky=tk.W, pady=5)
        ttk.Button(param_frame, text="浏览...", command=self._browse_path).grid(row=0, column=8, padx=5, pady=5)

        # 3. 图表合并配置
        merge_frame = ttk.LabelFrame(main_frame, text="图表合并配置", padding="10")
        merge_frame.pack(fill=tk.X, pady=(0, 10))

        self.merge_var = tk.BooleanVar(value=False)
        merge_check = ttk.Checkbutton(merge_frame, text="启用图表合并", variable=self.merge_var,
                                      command=self._toggle_merge_options)
        merge_check.grid(row=0, column=0, sticky=tk.W, padx=5, pady=5, columnspan=3)

        ttk.Label(merge_frame, text="当前监控进程:").grid(row=1, column=0, sticky=tk.W, padx=5)
        self.merge_source_listbox = tk.Listbox(merge_frame, selectmode=tk.EXTENDED, height=4, width=40)
        self.merge_source_listbox.grid(row=2, column=0, padx=5, pady=5)

        merge_btn_frame = ttk.Frame(merge_frame)
        merge_btn_frame.grid(row=2, column=1, padx=10)
        self.add_to_merge_btn = ttk.Button(merge_btn_frame, text="添加 >", command=self._add_to_merge,
                                           state=tk.DISABLED)
        self.add_to_merge_btn.pack(fill=tk.X, pady=5)
        self.remove_from_merge_btn = ttk.Button(merge_btn_frame, text="< 移除", command=self._remove_from_merge,
                                                state=tk.DISABLED)
        self.remove_from_merge_btn.pack(fill=tk.X, pady=5)

        ttk.Label(merge_frame, text="合并监控进程:").grid(row=1, column=2, sticky=tk.W, padx=5)
        self.merge_target_listbox = tk.Listbox(merge_frame, selectmode=tk.EXTENDED, height=4, width=40)
        self.merge_target_listbox.grid(row=2, column=2, padx=5, pady=5)

        # 4. 控制按钮区域
        control_frame = ttk.Frame(main_frame)
        control_frame.pack(fill=tk.X, pady=(0, 10))
        self.start_btn = ttk.Button(control_frame, text="开始监控", command=self._start_monitoring)
        self.start_btn.pack(side=tk.LEFT, padx=5)
        self.stop_btn = ttk.Button(control_frame, text="停止监控", command=self._stop_monitoring, state=tk.DISABLED)
        self.stop_btn.pack(side=tk.LEFT, padx=5)
        self.report_btn = ttk.Button(control_frame, text="生成报告", command=self._generate_report, state=tk.DISABLED)
        self.report_btn.pack(side=tk.LEFT, padx=5)
        self.status_var = tk.StringVar(value="就绪")
        ttk.Label(control_frame, textvariable=self.status_var).pack(side=tk.RIGHT, padx=5)

        # 5. 实时图表区域
        chart_frame = ttk.LabelFrame(main_frame, text="实时监控图表", padding="10")
        chart_frame.pack(fill=tk.BOTH, expand=True)  # 图表区域占满剩余空间

        self.fig, self.ax = plt.subplots(figsize=(10, 6))
        self.canvas = FigureCanvasTkAgg(self.fig, master=chart_frame)
        self.canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

        # 初始化悬停标注框
        self.annotation = self.ax.annotate(
            "",
            xy=(0, 0),
            xytext=(15, 15),
            textcoords="offset points",
            bbox=dict(boxstyle="round,pad=0.5", fc="white", ec="gray", alpha=0.9),
            arrowprops=dict(arrowstyle="->", connectionstyle="arc3,rad=0.2")
        )
        self.annotation.set_visible(False)
        self.canvas.mpl_connect("motion_notify_event", self._on_mouse_hover)

        # ---------------------- 修改2：调整统计信息模块位置（上移以完全显示） ----------------------
        stats_frame = ttk.LabelFrame(main_frame, text="统计信息", padding="10")
        # 关键调整：减少上方边距，取消side=tk.BOTTOM，通过pack顺序自然位于图表下方
        stats_frame.pack(fill=tk.X, pady=(5, 0))

        # 创建统计数据表格（Treeview控件）并设置居中显示
        columns = ("proc", "max", "min", "avg", "3sigma")
        self.stats_tree = ttk.Treeview(stats_frame, columns=columns, show="headings", height=6)  # 增加height为6行

        # 设置表头
        self.stats_tree.heading("proc", text="进程名")
        self.stats_tree.heading("max", text="最大值 (MB)")
        self.stats_tree.heading("min", text="最小值 (MB)")
        self.stats_tree.heading("avg", text="平均值 (MB)")
        self.stats_tree.heading("3sigma", text="3σ值 (MB)")

        # 设置列宽和居中对齐
        self.stats_tree.column("proc", width=150, anchor="center")
        self.stats_tree.column("max", width=100, anchor="center")
        self.stats_tree.column("min", width=100, anchor="center")
        self.stats_tree.column("avg", width=100, anchor="center")
        self.stats_tree.column("3sigma", width=100, anchor="center")

        self.stats_tree.pack(fill=tk.X)

        self._refresh_processes()

    # 图表合并相关方法
    def _toggle_merge_options(self):
        state = tk.NORMAL if self.merge_var.get() else tk.DISABLED
        self.merge_source_listbox.config(state=state)
        self.merge_target_listbox.config(state=state)
        self.add_to_merge_btn.config(state=state)
        self.remove_from_merge_btn.config(state=state)
        if state == tk.NORMAL:
            self._sync_merge_source_list()

    def _sync_merge_source_list(self):
        self.merge_source_listbox.delete(0, tk.END)
        for i in range(self.monitor_listbox.size()):
            proc_name = self.monitor_listbox.get(i)
            self.merge_source_listbox.insert(tk.END, proc_name)

    def _add_to_merge(self):
        selected_indices = self.merge_source_listbox.curselection()
        for i in selected_indices:
            proc_name = self.merge_source_listbox.get(i)
            if proc_name not in [self.merge_target_listbox.get(j) for j in range(self.merge_target_listbox.size())]:
                self.merge_target_listbox.insert(tk.END, proc_name)

    def _remove_from_merge(self):
        selected_indices = self.merge_target_listbox.curselection()
        for i in sorted(selected_indices, reverse=True):
            self.merge_target_listbox.delete(i)

    # 进程管理相关方法
    def _refresh_processes(self):
        self.process_listbox.delete(0, tk.END)
        processes = set()
        for proc in psutil.process_iter(['name']):
            try:
                processes.add(proc.info['name'])
            except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
                continue
        for proc_name in sorted(processes):
            self.process_listbox.insert(tk.END, proc_name)

    def _add_monitor(self):
        selected_indices = self.process_listbox.curselection()
        for i in selected_indices:
            proc_name = self.process_listbox.get(i)
            if proc_name not in [self.monitor_listbox.get(j) for j in range(self.monitor_listbox.size())]:
                self.monitor_listbox.insert(tk.END, proc_name)
                if self.merge_var.get():
                    self._sync_merge_source_list()

    def _remove_monitor(self):
        selected_indices = self.monitor_listbox.curselection()
        for i in sorted(selected_indices, reverse=True):
            self.monitor_listbox.delete(i)
            if self.merge_var.get():
                self._sync_merge_source_list()

    def _browse_path(self):
        path = filedialog.askdirectory()
        if path:
            self.path_var.set(path)
            self.save_path = path

    # 监控控制方法
    def _start_monitoring(self):
        if self.monitor_listbox.size() == 0:
            messagebox.showwarning("警告", "请至少选择一个进程进行监控")
            return

        try:
            duration_value = int(self.duration_var.get())
            duration_unit = self.duration_unit.get()
            if duration_unit == "分钟":
                duration = duration_value * 60
            elif duration_unit == "小时":
                duration = duration_value * 3600
            else:
                duration = duration_value

            interval_value = int(self.interval_var.get())
            interval_unit = self.interval_unit.get()
            if interval_unit == "分钟":
                interval = interval_value * 60
            else:
                interval = interval_value

            if duration <= 0 or interval <= 0 or interval > duration:
                raise ValueError
        except ValueError:
            messagebox.showwarning("警告", "请输入有效的监控参数（正整数，且间隔不大于时长）")
            return

        self.process_data = {}
        for i in range(self.monitor_listbox.size()):
            proc_name = self.monitor_listbox.get(i)
            self.process_data[proc_name] = []

        self.status_var.set("监控中...")
        self.start_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)
        self.report_btn.config(state=tk.DISABLED)
        self.monitoring = True

        self.monitor_thread = threading.Thread(
            target=self._monitor_processes,
            args=(duration, interval),
            daemon=True
        )
        self.monitor_thread.start()

    def _stop_monitoring(self):
        self.monitoring = False
        self.status_var.set("监控已停止，准备生成报告")
        self.start_btn.config(state=tk.NORMAL)
        self.stop_btn.config(state=tk.DISABLED)
        self.report_btn.config(state=tk.NORMAL)

    def _monitor_processes(self, duration, interval):
        end_time = time.time() + duration
        while self.monitoring and time.time() < end_time:
            timestamp = datetime.now()
            for proc_name in self.process_data.keys():
                mem_usage = 0
                for proc in psutil.process_iter(['name', 'memory_info']):
                    try:
                        if proc.info['name'] == proc_name:
                            mem_usage += proc.info['memory_info'].rss
                    except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
                        continue
                self.process_data[proc_name].append({'Timestamp': timestamp, 'Memory_Bytes': mem_usage})
            self.root.after(0, self._update_chart)
            time.sleep(interval)
        if self.monitoring:
            self.root.after(0, lambda: self.status_var.set("监控完成，准备生成报告"))
            self.root.after(0, lambda: self.start_btn.config(state=tk.NORMAL))
            self.root.after(0, lambda: self.stop_btn.config(state=tk.DISABLED))
            self.root.after(0, lambda: self.report_btn.config(state=tk.NORMAL))
            self.monitoring = False

    # 更新统计信息表格
    def _update_stats_table(self, stats_data):
        """更新UI中的统计信息表格"""
        for item in self.stats_tree.get_children():
            self.stats_tree.delete(item)
        for data in stats_data:
            self.stats_tree.insert("", "end", values=data)

    # 更新图表
    def _update_chart(self):
        """更新实时图表"""
        self.ax.clear()
        has_data = False
        stats_data = []

        for proc_name, data in self.process_data.items():
            if data:
                has_data = True
                df = pd.DataFrame(data)
                df['Memory_MB'] = df['Memory_Bytes'] / (1024 * 1024)
                self.ax.plot(df['Timestamp'], df['Memory_MB'], marker='o', linestyle='-', label=proc_name)

                # 计算统计值
                memory_values = df['Memory_MB'].values
                max_val = np.max(memory_values) if len(memory_values) > 0 else 0
                min_val = np.min(memory_values) if len(memory_values) > 0 else 0
                avg_val = np.mean(memory_values) if len(memory_values) > 0 else 0
                std_val = np.std(memory_values, ddof=1) if len(memory_values) >= 2 else 0
                sigma3_val = 3 * std_val
                stats_data.append(
                    (proc_name, round(max_val, 2), round(min_val, 2), round(avg_val, 2), round(sigma3_val, 2)))

        # 更新统计表格
        if stats_data:
            self._update_stats_table(stats_data)

        # 图表样式设置
        if has_data:
            self.ax.set_title('实时内存使用监控')
            self.ax.set_xlabel('时间')
            self.ax.set_ylabel('内存使用 (MB)')
            self.ax.legend(loc='upper left')
            max_ticks = min(10, len(df))
            self.ax.xaxis.set_major_locator(plt.MaxNLocator(max_ticks))
            self.ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M:%S'))
            plt.xticks(rotation=45, ha='right')
            self.fig.tight_layout()

        self.canvas.draw()

    # 鼠标悬停事件
    def _on_mouse_hover(self, event):
        if event.inaxes != self.ax:
            self.annotation.set_visible(False)
            self.canvas.draw_idle()
            return
        if not self.ax.lines:
            self.annotation.set_visible(False)
            self.canvas.draw_idle()
            return
        mouse_x = event.xdata
        mouse_y = event.ydata
        if mouse_x is None or mouse_y is None:
            self.annotation.set_visible(False)
            self.canvas.draw_idle()
            return
        mouse_datetime = mdates.num2date(mouse_x)
        min_dist = float('inf')
        closest_point = None
        for line in self.ax.lines:
            proc_name = line.get_label()
            x_data = line.get_xdata()
            y_data = line.get_ydata()
            for i in range(len(x_data)):
                point_datetime = mdates.num2date(x_data[i])
                time_diff = abs((mouse_datetime - point_datetime).total_seconds())
                memory_diff = abs(mouse_y - y_data[i])
                if time_diff < 5 and memory_diff < 10:
                    distance = time_diff * 0.1 + memory_diff * 0.01
                    if distance < min_dist:
                        min_dist = distance
                        closest_point = {
                            "proc": proc_name,
                            "time": point_datetime.strftime("%Y-%m-%d %H:%M:%S"),
                            "memory": round(y_data[i], 2)
                        }
        if closest_point:
            self.annotation.xy = (mouse_x, mouse_y)
            self.annotation.set_text(
                f"进程: {closest_point['proc']}\n"
                f"时间: {closest_point['time']}\n"
                f"内存: {closest_point['memory']} MB"
            )
            self.annotation.set_visible(True)
        else:
            self.annotation.set_visible(False)
        self.canvas.draw_idle()

    # ---------------------- 修改3：完善统计信息保存至Excel ----------------------
    def _generate_report(self):
        """生成Excel报告（确保统计信息保存）"""
        if not self.process_data or all(len(data) == 0 for data in self.process_data.values()):
            messagebox.showwarning("警告", "没有监控数据可生成报告")
            return

        # 报告路径调整为"内存监控报告.xlsx"
        excel_path = os.path.join(self.save_path, "内存监控报告.xlsx")
        wb = Workbook()

        # 1. 创建内存数据工作表
        data_ws = wb.active
        data_ws.title = "内存监控数据"

        # 写入内存数据
        summary_data = {}
        process_names = [self.monitor_listbox.get(i) for i in range(self.monitor_listbox.size())]
        for proc_name in process_names:
            if proc_name in self.process_data and self.process_data[proc_name]:
                df = pd.DataFrame(self.process_data[proc_name])
                df['Memory_MB'] = df['Memory_Bytes'] / (1024 * 1024)
                df['Time'] = df['Timestamp'].dt.strftime('%H:%M:%S')

                for _, row in df.iterrows():
                    timestamp = row['Timestamp']
                    if timestamp not in summary_data:
                        summary_data[timestamp] = {'Timestamp': timestamp, 'Time': row['Time']}
                    summary_data[timestamp][proc_name] = row['Memory_MB']

        if summary_data:
            summary_df = pd.DataFrame.from_dict(summary_data, orient='index').sort_values('Timestamp')
            headers = ['时间戳'] + process_names
            data_ws.append(headers)

            for _, row in summary_df.iterrows():
                row_data = [row['Timestamp']] + [row.get(proc, '') for proc in process_names]
                data_ws.append(row_data)

            # 设置数据表格格式
            center_alignment = Alignment(horizontal='center', vertical='center')
            for row in data_ws.iter_rows(min_row=1, max_row=data_ws.max_row, min_col=1, max_col=data_ws.max_column):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0.00'
                    cell.alignment = center_alignment
            data_ws.column_dimensions['A'].width = 25
            for col in range(2, 2 + len(process_names)):
                data_ws.column_dimensions[chr(64 + col)].width = 18

        # 2. 创建统计汇总工作表（确保统计信息保存）
        stats_ws = wb.create_sheet(title="统计汇总")
        stats_ws.append(["进程名", "最大值 (MB)", "最小值 (MB)", "平均值 (MB)", "3σ值 (MB)"])

        # 计算并写入统计值
        for proc_name in process_names:
            if proc_name in self.process_data and self.process_data[proc_name]:
                df = pd.DataFrame(self.process_data[proc_name])
                df['Memory_MB'] = df['Memory_Bytes'] / (1024 * 1024)
                memory_values = df['Memory_MB'].values

                max_val = np.max(memory_values) if len(memory_values) > 0 else 0
                min_val = np.min(memory_values) if len(memory_values) > 0 else 0
                avg_val = np.mean(memory_values) if len(memory_values) > 0 else 0
                std_val = np.std(memory_values, ddof=1) if len(memory_values) >= 2 else 0  # 样本标准差
                sigma3_val = 3 * std_val

                stats_ws.append([
                    proc_name,
                    round(max_val, 2),
                    round(min_val, 2),
                    round(avg_val, 2),
                    round(sigma3_val, 2)
                ])

        # 设置统计表格格式
        for row in stats_ws.iter_rows(min_row=2, max_row=stats_ws.max_row, min_col=2, max_col=5):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')
        stats_ws.column_dimensions['A'].width = 15
        for col in ['B', 'C', 'D', 'E']:
            stats_ws.column_dimensions[col].width = 12

        # 生成图表
        current_row = len(summary_df) + 3 if summary_data else 1
        merge_chart_inserted = False
        if self.merge_var.get() and self.merge_target_listbox.size() > 0:
            merge_procs = [self.merge_target_listbox.get(i) for i in range(self.merge_target_listbox.size())]
            if merge_procs:
                fig, ax = plt.subplots(figsize=(12, 6))
                colors = ['blue', 'green', 'red', 'purple', 'orange', 'brown', 'pink', 'gray']
                color_idx = 0
                for proc_name in merge_procs:
                    if proc_name in self.process_data and self.process_data[proc_name]:
                        df = pd.DataFrame(self.process_data[proc_name])
                        df['Memory_MB'] = df['Memory_Bytes'] / (1024 * 1024)
                        ax.plot(df['Timestamp'], df['Memory_MB'], marker='o', linestyle='-',
                                label=proc_name, color=colors[color_idx % len(colors)])
                        color_idx += 1
                ax.set_title('多进程内存使用对比（合并图表）')
                ax.set_xlabel('时间')
                ax.set_ylabel('内存使用 (MB)')
                ax.legend()
                max_ticks = min(10, len(df)) if 'df' in locals() else 5
                ax.xaxis.set_major_locator(plt.MaxNLocator(max_ticks))
                ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M:%S'))
                plt.xticks(rotation=45, ha='right')
                plt.tight_layout()
                img_data = BytesIO()
                fig.savefig(img_data, format='png')
                img_data.seek(0)
                img = Image(img_data)
                img.width = 800
                img.height = 400
                data_ws.add_image(img, f'A{current_row}')
                current_row += 30
                plt.close(fig)
                merge_chart_inserted = True

        if not merge_chart_inserted:
            current_row = len(summary_df) + 3 if summary_data else 1
        else:
            current_row += 5

        for proc_name in process_names:
            if proc_name in self.process_data and self.process_data[proc_name]:
                df = pd.DataFrame(self.process_data[proc_name])
                df['Memory_MB'] = df['Memory_Bytes'] / (1024 * 1024)
                fig, ax = plt.subplots(figsize=(10, 4))
                ax.plot(df['Timestamp'], df['Memory_MB'], marker='o', linestyle='-', color='blue')
                ax.set_title(f'{proc_name} 内存使用趋势')
                ax.set_xlabel('时间')
                ax.set_ylabel('内存使用 (MB)')
                max_ticks = min(10, len(df))
                ax.xaxis.set_major_locator(plt.MaxNLocator(max_ticks))
                ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M:%S'))
                plt.xticks(rotation=45, ha='right')
                plt.tight_layout()
                img_data = BytesIO()
                fig.savefig(img_data, format='png')
                img_data.seek(0)
                img = Image(img_data)
                img.width = 600
                img.height = 300
                data_ws.add_image(img, f'A{current_row}')
                current_row += 20
                plt.close(fig)

        try:
            wb.save(excel_path)
            messagebox.showinfo("成功", f"报告已生成：\n{excel_path}")
            self.status_var.set("报告生成完成")
        except Exception as e:
            messagebox.showerror("错误", f"保存报告失败：{str(e)}")


if __name__ == "__main__":
    root = tk.Tk()

    # ---------------------- 添加图标设置代码 ----------------------
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))    # 获取脚本所在目录的绝对路径
        icon_path = os.path.join(script_dir, "app_icon.ico")    # 拼接图标文件路径（确保图标文件与脚本同目录）
        root.iconbitmap(icon_path)    # 设置窗口图标（同时影响任务栏图标）
    except Exception as e:
        # 图标加载失败时不影响程序运行，仅打印提示
        print(f"图标设置失败：{str(e)}（请确保app_icon.ico文件存在于脚本目录）")

    app = MemoryMonitorApp(root)
    root.mainloop()